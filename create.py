import pandas
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = load_workbook('Grades.xlsx')

# modify sheet
# ws = wb.active
# ws['A1'].value = "Test"
# print(ws['A1'].value)
# wb.save('Grades.xlsx')

# create sheet
# wb.create_sheet('Test')
# print(wb.sheetnames)

# create new workbook
# wb = Workbook()
# ws = wb.active
# ws.title = "Data"

# ws.append(['Tim', 'Is', 'Great', '!'])
# ws.append(['Sam', 'Is', 'Great', '!'])
# ws.append(['John', 'Is', 'Great', '!'])
# ws.append(['Mimi', 'Is', 'Great', '!'])
# wb.save('mimi.xlsx')


# accessing multiple cells
# wb = load_workbook('mimi.xlsx')
# ws = wb.active

# for row in range(1, 11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         ws[char + str(row)] = char + str(row)
#         print(ws[char + str(row)].value)

# wb.save('mimi.xlsx')

# merging cells
# wb = load_workbook('mimi.xlsx')
# ws = wb.active

# ws.merge_cells("A1:D2")
# ws.unmerge_cells("A1:D1")
# wb.save("mimi.xlsx")

# merging cells
# wb = load_workbook('mimi.xlsx')
# ws = wb.active

# ws.insert_rows(7)
# ws.insert_rows(7)

# ws.move_range("C1:D11", rows=2, cols=2)
# wb.save("mimi.xlsx")