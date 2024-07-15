# from openpyxl import Workbook, load_workbook
import openpyxl
import pandas as pd
import requests


global wb

def createNewFile(new_name):
    global Workbook

    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]

    sheet['C3'] = 'Hello World'

    wb.create_sheet('New Sheet')
    wb.save(new_name+ '.xlsx')

def loadOld(filename_with_direc):
    global wb

    wb = openpyxl.load_workbook(filename_with_direc, read_only=True)

    def pickWantedSheet():

        loadOld(filename_with_direc)

        print(wb.sheetnames)
        name = input('Input the name you want--->')
        sheet = wb[name]

        def CreateNewSheet():

            loadOld(filename_with_direc)
            wb.create_sheet(input('Input A Name--->'))
            wb.save(filename_with_direc)

            loadOld(filename_with_direc)

            for row in sheet.rows:
                for cell in row:
                    print(cell.value)

            for i in range(sheet.max_row):
                for j in range(sheet.max_column):
                    print(sheet.cell(row=i+1, column=j+1).value)
# response = requests.get("https://api.coincap.io/v2/assets")
# data = response.json()

# dataF = []
# for x in data['data']:
#     # date = x['date']
#     # for s in x['sources']:
#         dataF.append({
#             # "Date": date,
#             "Source": x['name'],
#             "Forecast": x['symbol'],
#             "Actual": x['id']
#         })

# df = pd.DataFrame(dataF)

# excel_file = 'mich.xlsx'
# sheet_name = 'Sheet1'

# try:
#     book = load_workbook(excel_file)
#     writer = pd.ExcelWriter(excel_file, engine='openpyxl')
#     writer.book = book
# except FileNotFoundError:
#     writer = pd.ExcelWriter(excel_file, engine='openpyxl')
#     book = Workbook()
#     # writer.book =book

# df.to_excel(writer, sheet_name=sheet_name, index=False)

# writer.save()
# writer.close()

# print("Data mimi is here", excel_file)