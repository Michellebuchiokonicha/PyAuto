from openpyxl import Workbook, load_workbook
import requests
from datetime import datetime, timedelta

import schedule
import time

api_url = "https://opendata.elia.be/api/explore/v2.1/catalog/datasets/ods087/records?limit=100&refine=region%3A%22Belgium%22"
excel_file = "Price_forecast_Day_Ahead.xlsx"

def fetch_energy_data(offset=0):
    response = requests.get(api_url + f"&offset={offset}")
    data = response.json()

    if response.status_code == 200:
        data = response.json()
        return data["results"], data["total_count"] 
    else:
        print(f"Error fetching data: {response.status_code}")
        return [], 0

def update_excel_data(data):
    try:
        wb = load_workbook(excel_file)
        ws = wb.worksheets[0]  

        for row in range(5, ws.max_row + 1):  
            for col in range(1, 9):  
                ws.cell(row=row, column=col).value = None  

                now = datetime.now()
                current_year = now.year
                current_month = now.month

        start_date = datetime(current_year,current_month, 1) 
        end_date = datetime(current_year, current_month, 24) 

        filtered_data = [
            result
            for result in data
            if start_date <= datetime.fromisoformat(result["datetime"]).replace(tzinfo=None) <= end_date
        ]

      
        for i, result in enumerate(filtered_data):  
            row = i + 5  
            ws[f"A{row}"] = result["datetime"]
            ws[f"B{row}"] = result["mostrecentforecast"]
            ws[f"C{row}"] = result["dayaheadforecast"]
            ws[f"D{row}"] = result["weekaheadforecast"]
            ws[f"E{row}"] = result["realtime"]
            ws[f"F{row}"] = result["realtime"]  
            ws[f"G{row}"] = result["monitoredcapacity"]
            ws[f"H{row}"] = result["dayahead11hforecast"]

        for row in range(5, ws.max_row + 1):
            ws[f"I{row}"] = ws[f"I{row}"].value  
            ws[f"J{row}"] = ws[f"J{row}"].value  
            ws[f"K{row}"] = ws[f"K{row}"].value  
            ws[f"L{row}"] = ws[f"L{row}"].value  

        wb.save(excel_file)
        print(f"Excel file updated: {excel_file}")
    except FileNotFoundError:
        print(f"Excel file not found: {excel_file}")
    except KeyError:
        print(f"Sheet 'Forecast PV' not found in the Excel file.")
    schedule.every().hour.do(update_excel_data)

    while True:
             schedule.run_pending()

if __name__ == "__main__":
    all_data = []
    offset = 0
    total_count = 0
    while True:
        data, total_count = fetch_energy_data(offset)  
        if not data:
            break
        all_data.extend(data)
        offset += 100  
        if offset >= total_count:  
            break

    update_excel_data(all_data)
