from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Joe" : {
        "math":65,
        "science": 78,
        "english": 98,
        "gym": 89
    },
    "Bill" : {
        "math":55,
        "science": 72,
        "english": 88,
        "gym": 77
    },
    "Tim" : {
        "math":100,
        "science": 66,
        "english": 93,
        "gym": 74
    },
    "Peace" : {
        "math":77,
        "science": 83,
        "english": 59,
        "gym": 91
    },
}


wb = Workbook()
ws = wb.active
ws.title = "Grades"
headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

for col in range(2, len(data['Joe']) + 2):
    char = get_column_letter(col)
    ws[char + '7'] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")


wb.save("NewGrades.xlsx")