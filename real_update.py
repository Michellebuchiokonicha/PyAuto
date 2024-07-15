from openpyxl import Workbook, load_workbook
import requests
from datetime import datetime, timedelta

api_url = "https://opendata.elia.be/api/explore/v2.1/catalog/datasets/ods087/records?limit=100&refine=region%3A%22Belgium%22"
excel_file = "Price_forecast_Day_Ahead.xlsx"

def fetch_energy_data(offset=0):
    response = requests.get(api_url + f"&offset={offset}")

    if response.status_code == 200:
        data = response.json()
        return data["results"]
    else:
        print(f"Error fetching data: {response.status_code}")
        return []

def update_excel_data(data):
    try:
        wb = load_workbook(excel_file)
        ws = wb.worksheets[0]  

        
        for row in range(5, ws.max_row + 1):  
            for col in range(1, 9):  
                ws.cell(row=row, column=col).value = None  

     
        start_date = datetime(2024, 7, 1) 
        end_date = datetime(2024, 7, 22) 

        filtered_data = [
            result
            for result in data
            if start_date <= datetime.fromisoformat(result["datetime"]).replace(tzinfo=None) <= end_date
        ]

       
        for i, result in enumerate(filtered_data):  
            row = i + 5  
            ws[f"A{row}"] = result["datetime"]
            ws[f"B{row}"] = result["mostrecentforecast"]
            ws[f"C{row}"] = result["dayaheadforecast"]
            ws[f"D{row}"] = result["weekaheadforecast"]
            ws[f"E{row}"] = result["realtime"]
            ws[f"F{row}"] = result["realtime"]  
            ws[f"G{row}"] = result["monitoredcapacity"]
            ws[f"H{row}"] = result["dayahead11hforecast"]

        wb.save(excel_file)
        print(f"Excel file updated: {excel_file}")
    except FileNotFoundError:
        print(f"Excel file not found: {excel_file}")
    except KeyError:
        print(f"Sheet 'Forecast PV June' not found in the Excel file.")

if __name__ == "__main__":
    all_data = []
    offset = 0
    while True:
        data = fetch_energy_data(offset)
        if not data:
            break
        all_data.extend(data)
        offset += 100  

    update_excel_data(all_data)
