from openpyxl import Workbook, load_workbook
import requests
from datetime import datetime, timedelta

api_url = "https://opendata.elia.be/api/explore/v2.1/catalog/datasets/ods086/records?limit=100"
excel_file = "mimi_wind.xlsx"

# response = requests.get(api_url)
# if response.status_code == 200:
#     data = response.json()

# wb = Workbook()
# ws = wb.active
# ws.title = "Energy Consumption Data"

# ws.append(["Datetime", "Region", "Monitored Capacity"])

# for result in data["results"]:
#     ws.append([result["datetime"], result["region"], result["mostrecentforecast"]])

#     wb.save("mimi_wind.xlsx")

#     print("Excel file created: energy_data.xlsx")
# else:
#     print(f"Error fetching data:  {response.status_code}")

# end and start
def fetch_energy_data(offset=0):
    response = requests.get(api_url + f"&offset={offset}")

    if response.status_code == 200:
        data = response.json()
        # print(data["total_count"])
        return data["results"]
    else:
        print(f"Error fetching data: {response.status_code}")
        return []

def update_excel_data(data):
    try:
        wb = load_workbook(excel_file)
        ws = wb.worksheets[0]  

        for row in range(7, ws.max_row + 1):  
            for col in range(1, 9):  
                ws.cell(row=row, column=col).value = None  

                now = datetime.now()
                current_year = now.year
                current_month = now.month

        start_date = datetime(current_year, current_month, 1) 
        end_date = datetime(current_year, current_month, 22) 

        filtered_data = [
            result
            for result in data
            if start_date <= datetime.fromisoformat(result["datetime"]).replace(tzinfo=None) <= end_date
        ]

      
        for i, result in enumerate(filtered_data):  
            row = i + 7
            ws[f"A{row}"] = result["datetime"]
            ws[f"B{row}"] = result["weekaheadforecast"]
            ws[f"C{row}"] = result["dayaheadforecast"]
            ws[f"D{row}"] = result["mostrecentforecast"]
            ws[f"E{row}"] = result["realtime"]
            ws[f"F{row}"] = result["monitoredcapacity"]  
            ws[f"G{row}"] = result["decrementalbidid"]
            ws[f"H{row}"] = result["dayahead11hforecast"]

        for row in range(7, ws.max_row + 1):
            ws[f"I{row}"] = ws[f"I{row}"].value  
            ws[f"J{row}"] = ws[f"J{row}"].value  
            ws[f"K{row}"] = ws[f"K{row}"].value  
            ws[f"L{row}"] = ws[f"L{row}"].value

        wb.save(excel_file)
        print(f"Excel file updated: {excel_file}")
    except FileNotFoundError:
        print(f"Excel file not found: {excel_file}")
    except KeyError:
        print(f"Sheet 'Forecast wind generation' not found in the Excel file.")

if __name__ == "__main__":
    all_data = []
    offset = 0
    while True:
        data = fetch_energy_data(offset)
        if not data:
            break
        all_data.extend(data)
        offset += 100  

    update_excel_data(all_data)